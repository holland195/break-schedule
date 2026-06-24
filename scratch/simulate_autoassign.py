import json
import re
import sys
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8')

# 1. Load data
with open(r"C:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\scratch\current_users.json", "r", encoding="utf-8") as f:
    current_users = json.load(f)

with open(r"C:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\scratch\extracted_jul_26_correct.json", "r", encoding="utf-8") as f:
    jul_users = json.load(f)

# Load full schedule from excel for the week of June 29
import openpyxl
file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb["Schedule Jul_26"]

# Date columns in July sheet
date_cols = {}
for c in range(9, sheet.max_column + 1):
    val = sheet.cell(row=1, column=c).value
    if isinstance(val, str):
        date_cols[val] = c
    elif val:
        date_cols[val.strftime("%d/%m")] = c

# We want the week of June 29: 29/06, 30/06, 01/07, 02/07, 03/07, 04/07, 05/07
week_dates = ["29/06", "30/06", "01/07", "02/07", "03/07", "04/07", "05/07"]
print("Week dates of June 29 in July sheet:", [d for d in week_dates if d in date_cols])

# Build schedule map: username -> date -> shiftCode
jul_users_map = {u['username']: u for u in jul_users}
schedule_data = {}
for u in jul_users:
    r = u['row']
    un = u['username']
    schedule_data[un] = {}
    for d in week_dates:
        if d in date_cols:
            val = sheet.cell(row=r, column=date_cols[d]).value
            val_str = str(val or '0').strip().upper()
            if val_str == 'OFF' or not val_str:
                val_str = '0'
            schedule_data[un][d] = val_str

# 2. Simulate cleaning corrupted users and updating valid users
valid_users = []
for u in current_users:
    username = u.get('username', '')
    is_corrupted = False
    if ' ' in username or any(c.isupper() for c in username) or any(ord(c) > 127 for c in username) or username in ["start", "agent", "qa"] or ":" in username:
        is_corrupted = True
    if not is_corrupted:
        valid_users.append(u)

# Update teams
jul_team_map = {u['username']: u['team'] for u in jul_users}
updated_users = []
for u in valid_users:
    un = u['username'].lower().strip()
    if un in jul_team_map:
        u['team'] = jul_team_map[un]
    updated_users.append(u)

# 3. Simulate auto-assign breaks rotation for June 29 week
# We'll use E1/E2 for E, A1/A2 for A, D1/D2 for D
# Split percentages: default 50%
def role_tier(role):
    r = role.lower().strip()
    if r in ['data analyst', 'sr data analyst', 'agent', 'sr agent']:
        return 'agent'
    if r in ['data supervisor', 'qa']:
        return 'qa'
    if r in ['sr data supervisor', 'sr qa']:
        return 'sr_qa'
    return None

# Group users by shift & tier
# For the week of June 29:
# A user is on a shift in the week if they have at least one day on that shift
shift_tier_users = {}
for u in updated_users:
    un = u['username']
    role = u.get('role', '')
    tier = role_tier(role)
    if not tier:
        continue
    
    # Check what shifts they are scheduled for during the week
    for d in week_dates:
        shift = schedule_data.get(un, {}).get(d, '0')
        if shift in ['A', 'D', 'E']:
            key = (shift, tier)
            if key not in shift_tier_users:
                shift_tier_users[key] = []
            if u not in shift_tier_users[key]:
                shift_tier_users[key].append(u)

# Run rotation simulation for each shift & tier
rotation_assignments = {}
for (shift, tier), members in shift_tier_users.items():
    # Group members by team
    team_groups = {}
    for u in members:
        t = u.get('team', '_no_team_')
        if t not in team_groups:
            team_groups[t] = []
        team_groups[t].append(u)
    
    teams = sorted(list(team_groups.keys()))
    
    # 50/50 split
    slot1_count = (len(teams) + 1) // 2
    slot2_count = len(teams) - slot1_count
    
    # Simulate team slot assignment (Phase 0 since they are brand new teams)
    # The first slot2_count teams get slot2, rest get slot1
    # Note: _getTeamSlotMap assigns brandNew to slot2 first
    team_slots = {}
    for i, t in enumerate(teams):
        team_slots[t] = shift + '2' if i < slot2_count else shift + '1'
        
    rotation_assignments[(shift, tier)] = {
        "teams": teams,
        "team_slots": team_slots,
        "member_count": len(members)
    }

# Print summary
print("\n=== Break Assignment Simulation for Week of June 29 ===")
for (shift, tier), info in sorted(rotation_assignments.items()):
    print(f"\nShift {shift} | Tier {tier} | Total Members: {info['member_count']} | Total Teams: {len(info['teams'])}")
    print(f"  Slot 1 Teams: {len([t for t, s in info['team_slots'].items() if s.endswith('1')])}")
    print(f"  Slot 2 Teams: {len([t for t, s in info['team_slots'].items() if s.endswith('2')])}")
    # Print team allocations
    slot1_teams = sorted([t for t, s in info['team_slots'].items() if s.endswith('1')])
    slot2_teams = sorted([t for t, s in info['team_slots'].items() if s.endswith('2')])
    print(f"  Slot 1 Teams list: {slot1_teams[:10]} ...")
    print(f"  Slot 2 Teams list: {slot2_teams[:10]} ...")
