import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in ["Schedule Jun_26", "Schedule Jul_26"]:
    if sheet_name in wb.sheetnames:
        print(f"\n=== First 3 Rows of {sheet_name} ===")
        sheet = wb[sheet_name]
        for r in range(1, 4):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
            print(f"Row {r}: {row_vals}")
