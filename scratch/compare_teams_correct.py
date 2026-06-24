import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb["Schedule Jul_26"]

records = []
max_row = sheet.max_row

# Header row is row 1
header = [cell.value for cell in sheet[1]]

# Data starts at row 4
for r in range(4, max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
    if not any(row_vals):
        continue
    
    # Correct columns for Schedule Jul_26:
    # Col B is 2 (index 1): Team (new group)
    # Col C is 3 (index 2): Old Team
    # Col D is 4 (index 3): Name
    # Col E is 5 (index 4): Username
    # Col F is 6 (index 5): EmpNo
    # Col G is 7 (index 6): Role
    
    team = row_vals[1]
    old_team = row_vals[2]
    name = row_vals[3]
    username = row_vals[4]
    emp_no = row_vals[5]
    role = row_vals[6]
    
    if username and not isinstance(username, (int, float)):
        username_str = str(username).strip().lower()
        if username_str and not username_str.startswith("username") and username_str not in ["start", "agent", "qa"]:
            records.append({
                "row": r,
                "team": str(team).strip() if team is not None else "",
                "old_team": str(old_team).strip() if old_team is not None else "",
                "name": str(name).strip() if name is not None else "",
                "username": username_str,
                "empNo": str(emp_no).strip() if emp_no is not None else "",
                "role": str(role).strip() if role is not None else ""
            })

print(f"Found {len(records)} valid user records in Schedule Jul_26.")

# Save to a scratch JSON file for correctness
out_path = r"C:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\scratch\extracted_jul_26_correct.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(records, f, indent=2, ensure_ascii=False)
print(f"Saved correct records to {out_path}")

# Load current Firebase users
with open(r"C:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\scratch\current_users.json", "r", encoding="utf-8") as f:
    current_users = json.load(f)

# Compare
current_map = {u['username'].lower().strip(): u for u in current_users if 'username' in u}
jul_map = {u['username']: u for u in records}

differences = []
added = []
removed = []

for username, jul_u in jul_map.items():
    if username in current_map:
        curr_u = current_map[username]
        curr_team = curr_u.get('team', '')
        jul_team = jul_u.get('team', '')
        if curr_team != jul_team:
            differences.append({
                "username": username,
                "name": curr_u.get('name', ''),
                "curr_team": curr_team,
                "new_team": jul_team
            })
    else:
        added.append(jul_u)

for username, curr_u in current_map.items():
    if username not in jul_map:
        removed.append(curr_u)

print(f"\nComparison:")
print(f"Differences in team assignment: {len(differences)}")
print(f"New users to add: {len(added)}")
print(f"Users in Firebase but not in Jul_26: {len(removed)}")

if differences:
    print("\n--- Team Assignment Changes (first 20) ---")
    for diff in differences[:20]:
        print(f"User: {diff['username']} ({diff['name']}) | Team: {diff['curr_team']} -> {diff['new_team']}")

if added:
    print("\n--- New Users to Add ---")
    for add in added:
        print(f"User: {add['username']} ({add['name']}) | Team: {add['team']} | Role: {add['role']}")
