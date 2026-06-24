import openpyxl
import json
import urllib.request
import sys
import time
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

url = 'https://break-schedule-pave-default-rtdb.asia-southeast1.firebasedatabase.app/bsched.json?auth=W0kg0YX5okfaQzWLFBiZwrY69WeK1YJufBQySZsK'

def is_corrupted_username(username):
    if not username:
        return True
    username = str(username).strip()
    if ' ' in username:
        return True
    if any(c.isupper() for c in username):
        return True
    if any(ord(c) > 127 for c in username):
        return True
    if username in ["start", "agent", "qa"]:
        return True
    if ":" in username:
        return True
    return False

def get_js_hash(s):
    h = 0
    for c in s:
        h = (31 * h + ord(c)) & 0xFFFFFFFF
    if h >= 0x80000000:
        h -= 0x100000000
    return abs(h)

def parse_date_header(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.strftime("%d/%m")
    if isinstance(val, str):
        val = val.strip()
        if '/' in val:
            parts = val.split('/')
            if len(parts) >= 2:
                return f"{parts[0].zfill(2)}/{parts[1].zfill(2)}"
    return None

try:
    # 1. Fetch current Firebase state
    print("Fetching current data from Firebase...")
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req) as response:
        res_data = response.read().decode('utf-8')
        wrapper = json.loads(res_data)
        
    if not wrapper or 'data' not in wrapper:
        print("Error: Invalid Firebase data format.")
        sys.exit(1)
        
    state = json.loads(wrapper['data'])
    print("Firebase state successfully loaded.")
    
    # 2. Backup current state to scratch/backup_state.json
    backup_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\scratch\backup_state.json"
    with open(backup_path, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, ensure_ascii=False)
    print(f"Saved database backup to {backup_path}")
    
    # 3. Clean corrupted users, staffInfo, and staffSchedule
    original_users = state.get('users', [])
    cleaned_users = []
    removed_users_cnt = 0
    
    for u in original_users:
        un = u.get('username', '')
        if is_corrupted_username(un):
            removed_users_cnt += 1
            continue
        cleaned_users.append(u)
        
    state['users'] = cleaned_users
    print(f"Cleaned users: Removed {removed_users_cnt} corrupted, {len(cleaned_users)} valid remaining.")
    
    original_staff_info = state.get('staffInfo', {})
    cleaned_staff_info = {}
    removed_staff_cnt = 0
    for uname, info in original_staff_info.items():
        if is_corrupted_username(uname):
            removed_staff_cnt += 1
            continue
        cleaned_staff_info[uname] = info
    state['staffInfo'] = cleaned_staff_info
    print(f"Cleaned staffInfo: Removed {removed_staff_cnt} corrupted, {len(cleaned_staff_info)} valid remaining.")
    
    original_sched = state.get('staffSchedule', {})
    cleaned_sched = {}
    removed_sched_cnt = 0
    for uname, sched in original_sched.items():
        if is_corrupted_username(uname):
            removed_sched_cnt += 1
            continue
        cleaned_sched[uname] = sched
    state['staffSchedule'] = cleaned_sched
    print(f"Cleaned staffSchedule: Removed {removed_sched_cnt} corrupted, {len(cleaned_sched)} valid remaining.")
    
    # 4. Read excel file and import correct records
    excel_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
    print(f"Loading Excel file {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb["Schedule Jul_26"]
    
    # Parse date columns starting at column 9 (I)
    date_cols = {}
    max_col = sheet.max_column
    for c in range(9, max_col + 1):
        cell_val = sheet.cell(row=1, column=c).value
        dk = parse_date_header(cell_val)
        if dk:
            date_cols[c] = dk
            
    print(f"Found {len(date_cols)} date columns in Excel: {list(date_cols.values())}")
    
    # Parse staff rows from row 4 to max_row
    max_row = sheet.max_row
    imported_cnt = 0
    updated_cnt = 0
    created_cnt = 0
    
    for r in range(4, max_row + 1):
        team = sheet.cell(row=r, column=2).value # Col B (NEW)
        name = sheet.cell(row=r, column=4).value # Col D (NAME)
        username = sheet.cell(row=r, column=5).value # Col E (USER)
        emp_no = sheet.cell(row=r, column=6).value # Col F (Emp No)
        role = sheet.cell(row=r, column=7).value # Col G (Role)
        
        if not username or not name:
            continue
            
        username_str = str(username).strip().lower()
        name_str = str(name).strip()
        team_str = str(team).strip() if team is not None else ""
        emp_no_str = str(emp_no).strip() if emp_no is not None else ""
        role_str = str(role).strip() if role is not None else ""
        
        if is_corrupted_username(username_str):
            print(f"Warning: Excel row {r} has corrupted username '{username_str}'. Skipping.")
            continue
            
        # Find in valid users
        user_obj = None
        for u in state['users']:
            if u.get('username') == username_str:
                user_obj = u
                break
                
        if user_obj:
            user_obj['team'] = team_str
            user_obj['role'] = role_str
            user_obj['name'] = name_str
            if emp_no_str:
                user_obj['empNo'] = emp_no_str
            updated_cnt += 1
        else:
            # Create user
            uid = get_js_hash(username_str)
            user_obj = {
                'id': uid,
                'username': username_str,
                'name': name_str,
                'team': team_str,
                'role': role_str,
                'gender': '',
                'empNo': emp_no_str
            }
            state['users'].append(user_obj)
            created_cnt += 1
            
        # Update staffInfo
        if username_str not in state['staffInfo']:
            state['staffInfo'][username_str] = {}
        state['staffInfo'][username_str].update({
            'name': name_str,
            'role': role_str,
            'team': team_str,
            'empNo': emp_no_str
        })
        if 'id' not in state['staffInfo'][username_str]:
            state['staffInfo'][username_str]['id'] = user_obj['id']
            
        # Update staffSchedule
        if username_str not in state['staffSchedule']:
            state['staffSchedule'][username_str] = {}
            
        for col_idx, dk in date_cols.items():
            shift_val = sheet.cell(row=r, column=col_idx).value
            if shift_val is not None:
                shift_str = str(shift_val).strip().upper()
                if shift_str in ['OFF', '0', ''] or not shift_str:
                    shift_str = '0'
                elif '.' in shift_str:
                    try:
                        shift_str = str(int(float(shift_str)))
                    except ValueError:
                        pass
                state['staffSchedule'][username_str][dk] = shift_str
            else:
                state['staffSchedule'][username_str][dk] = '0'
                
        imported_cnt += 1

    print(f"Excel import summary:")
    print(f"  Processed {imported_cnt} staff rows.")
    print(f"  Updated existing users: {updated_cnt}")
    print(f"  Created new users: {created_cnt}")
    
    # 5. Save updated state back to Firebase
    now_ms = int(time.time() * 1000)
    state['_usersUpdatedAt'] = now_ms
    state['_updated'] = now_ms
    
    print("Uploading updated state to Firebase...")
    payload = {"data": json.dumps(state, ensure_ascii=False)}
    payload_json = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    
    put_req = urllib.request.Request(
        url,
        data=payload_json,
        headers={'Content-Type': 'application/json', 'User-Agent': 'Mozilla/5.0'},
        method='PUT'
    )
    
    with urllib.request.urlopen(put_req) as put_res:
        print("Firebase upload complete. Response status:", put_res.status)
        
except Exception as e:
    print("Error during sync process:", e)
    sys.exit(1)
