import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in ["Schedule Jun_26", "Schedule Jul_26"]:
    if sheet_name in wb.sheetnames:
        print(f"\n--- Row 4 of {sheet_name} ---")
        sheet = wb[sheet_name]
        row4 = [sheet.cell(row=4, column=c).value for c in range(1, 15)]
        for i, val in enumerate(row4):
            print(f"Col {i+1} (Letter {openpyxl.utils.get_column_letter(i+1)}): {val} (type: {type(val)})")
