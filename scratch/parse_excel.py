import openpyxl
import json
import sys

# Set standard output encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb["Schedule Jul_26"]

records = []
max_row = sheet.max_row
print(f"Total rows in sheet: {max_row}")

# Header row is row 1
header = [cell.value for cell in sheet[1]]
print("Headers (first 10):", header[:10])

# Data starts at row 4
for r in range(4, max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
    if not any(row_vals):
        continue
    # Col B is 2 (index 1), Col C is 3 (index 2), Col D is 4 (index 3), Col G is 7 (index 6)
    team = row_vals[1]
    name = row_vals[2]
    username = row_vals[3]
    role = row_vals[6]
    
    if username:
        records.append({
            "row": r,
            "team": str(team).strip() if team is not None else "",
            "name": str(name).strip() if name is not None else "",
            "username": str(username).strip().lower(),
            "role": str(role).strip() if role is not None else ""
        })

print(f"Found {len(records)} records with usernames.")

# Save to a scratch JSON file for easy reading
out_path = r"C:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\scratch\extracted_jul_26.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(records, f, indent=2, ensure_ascii=False)

print(f"Saved records to {out_path}")
