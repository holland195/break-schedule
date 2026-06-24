import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheet_jun = wb["Schedule Jun_26"]
sheet_jul = wb["Schedule Jul_26"]

# Date columns in June: H (8) to N (14) represent 22/06 to 28/06
# Let's map date to colIndex for June
dates_jun = {}
for c in range(8, sheet_jun.max_column + 1):
    val = sheet_jun.cell(row=1, column=c).value
    if isinstance(val, str):
        dates_jun[val] = c
    elif val:
        # datetime
        dates_jun[val.strftime("%d/%m")] = c

# Date columns in July: I (9) onwards
dates_jul = {}
for c in range(9, sheet_jul.max_column + 1):
    val = sheet_jul.cell(row=1, column=c).value
    if isinstance(val, str):
        dates_jul[val] = c
    elif val:
        dates_jul[val.strftime("%d/%m")] = c

# Common dates (excluding June 22 which is not in July sheet)
common_dates = sorted(list(set(dates_jun.keys()) & set(dates_jul.keys())))
print("Common dates:", common_dates)

# Build user maps by username
# In June: username in Col D (4), shift start at 8
users_jun = {}
for r in range(4, sheet_jun.max_row + 1):
    un = sheet_jun.cell(row=r, column=4).value
    if un and isinstance(un, str):
        un = un.strip().lower()
        users_jun[un] = r

# In July: username in Col E (5), shift start at 9
users_jul = {}
for r in range(4, sheet_jul.max_row + 1):
    un = sheet_jul.cell(row=r, column=5).value
    if un and isinstance(un, str):
        users_jul[un] = un.strip().lower()
        users_jul[un] = r

diffs = []
for un in sorted(list(set(users_jun.keys()) & set(users_jul.keys()))):
    r_jun = users_jun[un]
    r_jul = users_jul[un]
    for d in common_dates:
        c_jun = dates_jun[d]
        c_jul = dates_jul[d]
        val_jun = str(sheet_jun.cell(row=r_jun, column=c_jun).value or '0').strip().upper()
        val_jul = str(sheet_jul.cell(row=r_jul, column=c_jul).value or '0').strip().upper()
        if val_jun == 'OFF': val_jun = '0'
        if val_jul == 'OFF': val_jul = '0'
        if val_jun != val_jul:
            diffs.append({
                "username": un,
                "date": d,
                "jun_shift": val_jun,
                "jul_shift": val_jul
            })

print(f"Total discrepancies in schedule for common dates: {len(diffs)}")
if diffs:
    print("\nDiscrepancies (first 10):")
    for diff in diffs[:10]:
        print(f"  User: {diff['username']} on {diff['date']} | June: {diff['jun_shift']} | July: {diff['jul_shift']}")
