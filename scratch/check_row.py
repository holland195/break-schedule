import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb["Schedule Jul_26"]

row4 = [sheet.cell(row=4, column=c).value for c in range(1, 15)]
for i, val in enumerate(row4):
    print(f"Col {i+1} (Letter {openpyxl.utils.get_column_letter(i+1)}): {val} (type: {type(val)})")
