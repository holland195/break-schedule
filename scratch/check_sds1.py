import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb["Schedule Jul_26"]

row90 = [sheet.cell(row=90, column=c).value for c in range(1, 15)]
for i, val in enumerate(row90):
    print(f"Col {i+1} ({openpyxl.utils.get_column_letter(i+1)}): {val} (type: {type(val)})")
