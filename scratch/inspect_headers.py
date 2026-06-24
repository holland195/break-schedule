import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\mcuon\OneDrive\Documents\GitHub\break-schedule\referance\Agent Schedule_I.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb["Schedule Jul_26"]

header = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
for idx, h in enumerate(header):
    if h is not None:
        print(f"Col {idx+1} ({openpyxl.utils.get_column_letter(idx+1)}): {h} (type: {type(h)})")
